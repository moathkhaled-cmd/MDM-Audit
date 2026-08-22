# ============================================================================
# ArabWheels AE — MDM Spec Audit (GitHub Actions daily runner)
# ============================================================================
# Same audit logic as the Colab/OpenRouter version, but every file read/write
# goes through the Drive API (via drive_utils.py) instead of a local mounted
# filesystem, since GitHub Actions runners don't have Drive mounted.
#
# Each run:
#   1. Downloads the CSV from Drive (the audited one if it already exists
#      there, otherwise the original).
#   2. Lists every PDF under your brochures folder (metadata only — no PDF
#      bytes downloaded yet) and re-runs matching fresh every time (cheap,
#      local, and keeps things correct across the Colab -> Actions switch).
#   3. Processes brochure groups until DAILY_REQUEST_CAP is hit or the queue
#      is empty, downloading only the PDFs it actually needs to read today.
#   4. Uploads the updated CSV, highlighted .xlsx, discrepancy/new-trim
#      reports, audit log, and progress snapshot back to the SAME Drive
#      folder, overwriting the previous version of each — so tomorrow's run
#      picks up exactly where today's left off, same as the Colab version.
# ============================================================================
# ArabWheels AE — MDM Spec Audit (GitHub Actions daily runner — Gemini)
# ============================================================================
# Every file read/write goes through the Drive API (via drive_utils.py)
# instead of a local mounted filesystem, since GitHub Actions runners don't
# have Drive mounted.
#
# Each run:
#   1. Downloads the CSV from Drive (the audited one if it already exists
#      there, otherwise the original).
#   2. Lists every PDF under your brochures folder (metadata only — no PDF
#      bytes downloaded yet) and re-runs matching fresh every time (cheap,
#      local, and keeps things correct across environment switches).
#   3. Processes brochure groups, then no-brochure vehicles via web search,
#      then a cross-check pass — each stopping cleanly once DAILY_REQUEST_CAP
#      or the relevant API key(s) are exhausted, resuming automatically on
#      the next scheduled run.
#   4. Uploads the updated CSV, highlighted .xlsx, discrepancy/new-trim
#      reports, audit log, and progress snapshot back to the SAME Drive
#      folder, overwriting the previous version of each.
#
# API KEYS (3 total, 2 roles):
#   GEMINI_API_KEY   -> dedicated to web-search/grounded audits (no-brochure
#                       fallback + the cross-check pass's search step)
#   GEMINI_API_KEY_2 -> primary key for brochure audits (the "usual" path)
#   GEMINI_API_KEY_3 -> optional automatic failover if key 2 hits its daily
#                       quota mid-run; the script switches over and keeps going
# ============================================================================

import os
import re
import io
import json
import time
import uuid
import random
import logging
import threading
from datetime import datetime, timezone

import pandas as pd
import pypdf
from google import genai
from google.genai import types
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import drive_utils

logging.getLogger("pypdf").setLevel(logging.ERROR)

# ============================== CONFIG =====================================

# Folder IDs from the Drive URL (…/folders/<THIS PART>). Set as GitHub repo
# "Variables" (not secrets — folder IDs aren't sensitive), read here as env vars.
GDRIVE_ROOT_FOLDER_ID = os.environ['GDRIVE_ROOT_FOLDER_ID']            # "AW AE MDM Project" folder
GDRIVE_BROCHURES_FOLDER_ID = os.environ['GDRIVE_BROCHURES_FOLDER_ID']  # extracted-PDFs folder

SOURCE_CSV_NAME = 'Arabwheels AE MDM data - MDM Version specs.csv'   # only used if no audited copy exists yet
OUTPUT_CSV_NAME = 'AUDITED_Arabwheels_MDM_data.csv'
OUTPUT_XLSX_HIGHLIGHTED_NAME = 'AUDITED_Arabwheels_MDM_data_HIGHLIGHTED.xlsx'
DISCREPANCY_DETAIL_NAME = 'Discrepancy_Detail.csv'
NEW_TRIMS_NAME = 'New_Trims_Found_In_Brochures.csv'
AUDIT_LOG_NAME = 'audit_log.jsonl'
SUMMARY_NAME = 'audit_summary.txt'
PROGRESS_NAME = 'audit_progress.json'
NOT_YET_AUDITED_NAME = 'Not_Yet_Audited.csv'

LOCAL_DIR = './run_data'
os.makedirs(LOCAL_DIR, exist_ok=True)
def local(name):
    return os.path.join(LOCAL_DIR, name)

def _read_key(name):
    return os.environ.get(name, '').strip()

# Three keys, two roles:
#   - GEMINI_API_KEY  : dedicated to web-search/grounded calls (the no-brochure
#     fallback audit + the cross-check pass's search step). Kept separate so
#     that path's usage never eats into the brochure-audit keys' quota.
#   - GEMINI_API_KEY_2, GEMINI_API_KEY_3 : the "usual" path -- every brochure
#     audit call, plus the structured-JSON finalization step that follows a
#     web search. Key 3 is an automatic failover: if key 2 hits ITS OWN daily
#     quota, the script switches to key 3 for the rest of the run instead of
#     stopping. GEMINI_API_KEY_3 is optional -- if unset, key 2 alone is used
#     (no failover, same behavior as before this feature was added).
WEB_SEARCH_API_KEY = _read_key('GEMINI_API_KEY')
if not WEB_SEARCH_API_KEY:
    raise RuntimeError("Set GEMINI_API_KEY (dedicated to web-search audits).")
web_search_client = genai.Client(api_key=WEB_SEARCH_API_KEY)

_brochure_keys = [k for k in [_read_key('GEMINI_API_KEY_2'), _read_key('GEMINI_API_KEY_3')] if k]
if not _brochure_keys:
    raise RuntimeError("Set at least GEMINI_API_KEY_2 (used for brochure audits).")
brochure_clients = [genai.Client(api_key=k) for k in _brochure_keys]

MODEL_NAME = os.environ.get('GEMINI_MODEL', 'gemini-3.6-flash')

MAX_TEXT_CHARS = 40_000
GROUP_BATCH_SIZE = 10
SAVE_EVERY_N_GROUPS = 5
MAX_RETRIES = 5

def get_int_env(name, default):
    """Reads an int env var, falling back to default if unset OR empty.

    GitHub Actions sets an env var to '' (not omitted) when it's assigned
    from a repo variable/secret that doesn't exist, e.g.
    `RATE_LIMIT_RPM: ${{ vars.RATE_LIMIT_RPM }}` with no such repo variable
    defined. In that case os.environ.get(name, default) still returns ''
    (the key IS present), so int() blows up. This checks for that case
    explicitly.
    """
    val = os.environ.get(name, '').strip()
    return int(val) if val else default


# Free-tier Gemini Flash is typically ~10-15 RPM depending on your account —
# check https://ai.google.dev/gemini-api/docs/rate-limits and set this to
# roughly 80% of your actual limit.
RATE_LIMIT_RPM = get_int_env('RATE_LIMIT_RPM', 10)
SLEEP_BETWEEN_REQUESTS = 60.0 / RATE_LIMIT_RPM

# This is what makes it a genuine daily partition: each GitHub Actions run
# does at most this many requests, then stops cleanly and picks back up on
# tomorrow's scheduled run. Gemini free tier is also capped per-day (varies
# by model/account) — set this to comfortably under your actual daily quota.
DAILY_REQUEST_CAP = get_int_env('DAILY_REQUEST_CAP', 100)

MAKE_COL, MODEL_COL, GEN_COL, VERSION_COL = (
    'manufacturer_name', 'model_name', 'generation_name', 'version_name'
)
ID_COL = 'version_spec_id'


BINARY_SPECS = [
    'air_conditioner', 'power_windows', 'power_door_locks', 'power_steering',
    'anti_lock_braking', 'traction_control', 'immobilizer', 'cup_holders',
    'rear_folding_seat', 'rear_wiper', 'alloy_wheels', 'tubeless_tyres',
    'central_locking', 'remote_boot', 'steering_adjustment', 'tachometer',
    'child_safety_locks', 'fog_lights', 'defroster', 'defogger', 'am_fm_radio',
    'cassette_player', 'cd_player', 'sun_roof', 'moon_roof', 'is_imported',
    'cool_box', 'dvd_player', 'cruise_control', 'keyless_entry', 'rear_ac_vents',
    'front_speakers', 'rear_speakers', 'usb_and_auxillary_cable', 'heated_seats',
    'steering_switches', 'front_camera', 'rear_camera', 'push_start',
    'hill_start_assist_control', 'isofix_child_seat_anchors', 'vehicle_stability_control'
]

NUMERIC_SPECS = [
    'overall_height', 'overall_length', 'overall_width', 'wheel_base',
    'ground_clearance', 'boot_space', 'kerb_weight', 'no_of_doors',
    'displacement', 'engine_power', 'torque', 'no_of_cylinders',
    'valves_per_cylinder', 'turning_radius', 'seating_capacity',
    'fuel_tank_capacity', 'mileage_high_way', 'mileage_city', 'mileage_overall',
    'max_speed', 'wheel_size', 'number_of_airbags', 'electric_motor_power',
    'charging_time', 'range', 'battery_capacity', 'compression_ratio',
]

TEXT_SPECS = [
    'valve_mechanism', 'cylinder_config', 'engine_type', 'fuel_system',
    'steering_type', 'power_assisted', 'transmission_type', 'gear_box',
    'wheel_type', 'tyre_size', 'suspensions', 'brakes', 'seat_material_type',
    'key_type', 'drive_train', 'battery_type', 'spare_tyre', 'spare_tyre_size',
    'transmission_technology', 'driving_modes',
]

NEW_ROW_COPY_COLS = [
    'make_id', 'model_id', 'gen_id', MAKE_COL, MODEL_COL, GEN_COL,
    'gen_launch_date', 'gen_close_date',
]

TERMINAL_STATUSES = {
    "Verified Accurate",
    "Flagged: Spec Discrepancy",
    "Flagged: Mismatched Brochure PDF",
    "Flagged: Missing Brochure PDF",
    "Flagged: Missing Make/Model in Sheet",
    "Flagged: Ambiguous Brochure Match",
    "Flagged: Image-Only or Unreadable PDF",
}
# These two are matching-time only (no LLM call) — safe to re-open if a
# brochure that didn't exist before now does.
REMATCHABLE_STATUSES = {"Flagged: Missing Brochure PDF", "Flagged: Ambiguous Brochure Match"}

FILENAME_NOISE_TOKENS = {
    'brochure', 'catalogue', 'catalog', 'spec', 'specs', 'specsheet', 'sheet',
    'product', 'compressed', 'final', 'pdf', 'en', 'uae', 'ksa', 'huge', 'pro',
    'v1', 'v2', 'v3', 'v4', 'new', 'update', 'updated', 'copy', 'draft',
}

_detail_lock = threading.Lock()
_new_rows_lock = threading.Lock()
_log_lock = threading.Lock()

spec_discrepancy_details = []
new_rows_from_brochures = []
_missing_trim_seen = set()
CHANGED_CELLS = {}

_daily_lock = threading.Lock()
_daily_count = 0

def daily_cap_try_consume():
    global _daily_count
    with _daily_lock:
        if _daily_count >= DAILY_REQUEST_CAP:
            return False
        _daily_count += 1
        return True

def daily_cap_exhausted():
    with _daily_lock:
        return _daily_count >= DAILY_REQUEST_CAP

# --- Per-key-pool quota tracking (separate from the DAILY_REQUEST_CAP budget
# above, which is a simple request-count safety ceiling regardless of which
# key served it). This tracks whether Google itself has actually rejected a
# key for exceeding ITS real daily quota, and drives automatic failover. ---

_active_brochure_key_lock = threading.Lock()
_active_brochure_key_index = 0
_brochure_keys_exhausted = False

def get_active_brochure_client():
    with _active_brochure_key_lock:
        return brochure_clients[_active_brochure_key_index], _active_brochure_key_index

def advance_brochure_key(failed_index):
    """Called when the key at failed_index just hit its real daily quota.
    Switches to the next configured key for all subsequent calls. Returns
    False once every brochure-audit key has been exhausted."""
    global _active_brochure_key_index, _brochure_keys_exhausted
    with _active_brochure_key_lock:
        if failed_index != _active_brochure_key_index:
            return True  # another call already advanced past this key
        if _active_brochure_key_index + 1 < len(brochure_clients):
            _active_brochure_key_index += 1
            print(f"[API key] Brochure-audit key {failed_index + 1} hit its daily quota — "
                  f"switching to key {_active_brochure_key_index + 1}.")
            return True
        _brochure_keys_exhausted = True
        print(f"[API key] All {len(brochure_clients)} brochure-audit key(s) have hit their daily quota.")
        return False

def brochure_keys_exhausted():
    return _brochure_keys_exhausted

_web_search_key_exhausted = False

def mark_web_search_key_exhausted():
    global _web_search_key_exhausted
    _web_search_key_exhausted = True
    print("[API key] The web-search key has hit its daily quota — web-search audits and the "
          "cross-check pass will stop for the rest of this run (brochure audits continue normally).")

def web_search_key_exhausted():
    return _web_search_key_exhausted

def should_stop_brochure_pass():
    return daily_cap_exhausted() or brochure_keys_exhausted()

def stop_reason_brochure():
    if brochure_keys_exhausted():
        return "All brochure-audit API keys have hit their daily quota"
    return f"Daily request cap ({DAILY_REQUEST_CAP}) reached"

def should_stop_web_search_pass():
    # This path needs BOTH the web-search key (for grounding) and a working
    # brochure-pool key (for the structured-extraction step that follows the
    # search) -- if either is gone, further attempts here would just fail
    # one at a time instead of stopping cleanly.
    return daily_cap_exhausted() or web_search_key_exhausted() or brochure_keys_exhausted()

def stop_reason_web_search():
    if web_search_key_exhausted():
        return "The web-search API key has hit its daily quota"
    if brochure_keys_exhausted():
        return "All brochure-audit API keys have hit their daily quota (needed to finalize web-search results)"
    return f"Daily request cap ({DAILY_REQUEST_CAP}) reached"

def values_differ(old_val, new_val):
    if pd.isna(old_val) and (new_val is None or str(new_val).strip() == ''):
        return False
    try:
        return float(old_val) != float(new_val)
    except (ValueError, TypeError):
        return str(old_val).strip().lower() != str(new_val).strip().lower()

def log_event(payload: dict):
    payload["ts"] = datetime.now(timezone.utc).isoformat()
    with _log_lock:
        with open(local(AUDIT_LOG_NAME), "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")

# ============================== DRIVE SETUP =================================

print("Authenticating to Google Drive...")
drive, service_account_email = drive_utils.get_drive_service()

print("Checking folder access...")
drive_utils.check_folder_access(drive, GDRIVE_ROOT_FOLDER_ID, "GDRIVE_ROOT_FOLDER_ID", service_account_email)
drive_utils.check_folder_access(drive, GDRIVE_BROCHURES_FOLDER_ID, "GDRIVE_BROCHURES_FOLDER_ID", service_account_email)

existing_output_id = drive_utils.find_file_id_by_name(drive, GDRIVE_ROOT_FOLDER_ID, OUTPUT_CSV_NAME)
if existing_output_id:
    print("Found existing audited CSV on Drive — resuming from it.")
    drive_utils.download_to_path(drive, existing_output_id, local(OUTPUT_CSV_NAME))
    df = pd.read_csv(local(OUTPUT_CSV_NAME), low_memory=False)
else:
    print("No audited CSV on Drive yet — starting from the original source CSV.")
    source_id = drive_utils.find_file_id_by_name(drive, GDRIVE_ROOT_FOLDER_ID, SOURCE_CSV_NAME)
    if not source_id:
        raise FileNotFoundError(f"Couldn't find '{SOURCE_CSV_NAME}' in the Drive root folder.")
    drive_utils.download_to_path(drive, source_id, local(SOURCE_CSV_NAME))
    df = pd.read_csv(local(SOURCE_CSV_NAME), low_memory=False)

for col in ['Brochure_File_Found', 'Matched_Brochure_Path', 'Accuracy_Status', 'Discrepancies_Flagged',
            'Match_Confidence', 'Audit_Source', 'Audit_Progress']:
    if col not in df.columns:
        df[col] = False if col == 'Brochure_File_Found' else ""

df['Discrepancies_Flagged'] = df['Discrepancies_Flagged'].astype('object')
df['Accuracy_Status'] = df['Accuracy_Status'].astype('object')
for spec in BINARY_SPECS + NUMERIC_SPECS + TEXT_SPECS:
    if spec in df.columns:
        df[spec] = df[spec].astype('object')

total_rows = len(df)
already_terminal = df['Accuracy_Status'].isin(TERMINAL_STATUSES).sum()
print(f"Sheet has {total_rows} total rows — {already_terminal} already have a final audit "
      f"status, {total_rows - already_terminal} still need one.")

# ============================ MATCH ENGINE ===================================
# Re-run every time (cheap, local, no API cost) rather than trusting a
# previous run's Matched_Brochure_Path, since that column may have been
# written by a different environment (e.g. a prior Colab run) with a
# different path format.

_word_re = re.compile(r'[a-z0-9]+')

def tokenize(s: str) -> list:
    return _word_re.findall(str(s).lower())

print("\nListing brochure PDFs on Drive...")
pdf_index = {}       # file_id -> set(tokens from its virtual path, both variants)
pdf_vpaths = {}       # file_id -> human-readable virtual path (for logs/filenames)
pdf_search_variants = {}  # file_id -> (original_clean, split_clean) text pair used
                          # for regex whole-word matching. 'split_clean' adds a
                          # boundary at every letter<->digit transition so a
                          # glued model code like 'Rich7' still matches the
                          # sheet's separated 'Rich 7' -- checked in ADDITION to
                          # the original text, not instead of it, since an
                          # already-separated code like 'H9-catalogue.pdf' must
                          # keep matching 'H9' as a single token too.
for file_id, vpath in drive_utils.list_pdfs_recursive(drive, GDRIVE_BROCHURES_FOLDER_ID):
    original_clean = re.sub(r'[^a-z0-9]', ' ', vpath.lower())
    split_clean = re.sub(r'(?<=[0-9])(?=[a-z])', ' ', original_clean)
    split_clean = re.sub(r'(?<=[a-z])(?=[0-9])', ' ', split_clean)
    pdf_index[file_id] = set(tokenize(original_clean)) | set(tokenize(split_clean))
    pdf_vpaths[file_id] = vpath
    pdf_search_variants[file_id] = (original_clean, split_clean)
print(f"Total PDFs found: {len(pdf_index)}")

def find_best_match(make, model, generation, start_year, end_year):
    make_str = str(make).lower().strip() if pd.notna(make) else ''
    model_str = str(model).lower().strip() if pd.notna(model) else ''

    if not make_str or not model_str:
        return None, "Flagged: Missing Make/Model in Sheet", None

    make_tokens = set(tokenize(make_str))
    model_tokens = set(tokenize(model_str)) - FILENAME_NOISE_TOKENS

    if not make_tokens or not model_tokens:
        return None, "Flagged: Missing Make/Model in Sheet", None

    def model_tokens_present(path_variants, path_tokens):
        for m_token in model_tokens:
            pattern = r'(?<![a-z0-9])' + re.escape(m_token) + r'(?![a-z0-9])'
            if not any(re.search(pattern, variant) for variant in path_variants):
                return False
        return True

    # --- Tier 1: make AND model tokens both present (order doesn't matter --
    # token search, not positional) ---
    candidates = []
    for file_id, path_tokens in pdf_index.items():
        path_variants = pdf_search_variants[file_id]
        if not make_tokens.issubset(path_tokens):
            continue
        if not model_tokens_present(path_variants, path_tokens):
            continue
        gen_tokens = set(tokenize(generation)) if generation and str(generation).lower() != 'nan' else set()
        gen_hits = len(gen_tokens & path_tokens)
        score = (gen_hits, -len(path_tokens))
        candidates.append((score, file_id))

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1], None, "Exact"

    # --- Tier 2: many brochure filenames omit the manufacturer entirely
    # (e.g. 'H9-catalogue-UAE-EN-compressed.pdf', 'Rich7.pdf', 'Dargo.pdf').
    # Fall back to matching on model tokens alone. If more than one distinct
    # brochure matches, we can't safely guess which manufacturer it belongs
    # to (model codes like H9/S7/T5 get reused across different Chinese-
    # market brands) -- flag it for a human instead of picking one. ---
    tier2_candidates = []
    for file_id, path_tokens in pdf_index.items():
        path_variants = pdf_search_variants[file_id]
        if not model_tokens_present(path_variants, path_tokens):
            continue
        tier2_candidates.append(file_id)

    if len(tier2_candidates) == 1:
        return tier2_candidates[0], None, "Model-Only"
    if len(tier2_candidates) > 1:
        return None, "Flagged: Ambiguous Brochure Match", None

    return None, "Flagged: Missing Brochure PDF", None

print("Matching rows against brochures...")
rematched_count = 0
for index, row in df.iterrows():
    prior_status = row.get('Accuracy_Status')
    result, note, confidence = find_best_match(
        row.get(MAKE_COL, ''), row.get(MODEL_COL, ''), row.get(GEN_COL, ''),
        row.get('start_year'), row.get('end_year')
    )

    if note:
        df.at[index, 'Brochure_File_Found'] = False
        # Only overwrite a matching-time status; never touch an LLM-derived one.
        if prior_status not in TERMINAL_STATUSES or prior_status in REMATCHABLE_STATUSES:
            df.at[index, 'Accuracy_Status'] = note
    else:
        df.at[index, 'Brochure_File_Found'] = True
        df.at[index, 'Matched_Brochure_Path'] = pdf_vpaths[result]  # human-readable, for the sheet
        df.at[index, 'Match_Confidence'] = confidence
        if prior_status in REMATCHABLE_STATUSES:
            df.at[index, 'Accuracy_Status'] = ""  # newly matchable — reopen for audit
            rematched_count += 1

# lookup: vpath -> file_id, since Matched_Brochure_Path stores the readable
# path but downloading needs the Drive file id
vpath_to_id = {v: k for k, v in pdf_vpaths.items()}

if rematched_count:
    print(f"{rematched_count} row(s) that previously had no brochure now match one "
          f"(new PDFs added since the last run) — reopened for audit.")

# ============================ PDF TEXT (lazy download + cache) ==============

_pdf_text_cache = {}

def extract_pdf_text(vpath: str) -> str:
    if vpath in _pdf_text_cache:
        return _pdf_text_cache[vpath]
    file_id = vpath_to_id.get(vpath)
    if not file_id:
        return ""
    local_pdf_path = local(f"pdf_cache_{file_id}.pdf")
    try:
        drive_utils.download_to_path(drive, file_id, local_pdf_path)
        reader = pypdf.PdfReader(local_pdf_path)
        parts = []
        for i, page in enumerate(reader.pages):
            parts.append(f"\n--- PAGE {i + 1} ---\n" + (page.extract_text() or ""))
        full_text = "".join(parts)
        if len(full_text) > MAX_TEXT_CHARS:
            half = MAX_TEXT_CHARS // 2
            full_text = full_text[:half] + "\n\n...[TRUNCATED]...\n\n" + full_text[-half:]
    except Exception as e:
        print(f"  ! Failed to read/download {vpath}: {e}")
        full_text = ""
    finally:
        if os.path.exists(local_pdf_path):
            os.remove(local_pdf_path)  # runner disk is limited; don't hoard PDFs across the run
    _pdf_text_cache[vpath] = full_text
    return full_text

# ============================ GEMINI API CALL ================================

SYSTEM_INSTRUCTION = """
You are an automotive spec auditor. Be precise and literal. Never guess, assume, or invent a
value that isn't stated in the source text given to you.

Do not trust the current flags/specs given to you as correct — some are known to be wrong.
Re-derive every binary flag and spec value yourself from the source text only.

For every binary flag, reply with EXACTLY one of these three characters, nothing else:
  "1" = the source text explicitly confirms this trim HAS this feature.
  "0" = the source text explicitly confirms this trim does NOT have this feature
        (states it as unavailable, not fitted, "-", or "N/A" for this specific trim).
  "U" = the source text does not clearly state either way for this specific trim.
        Use "U" whenever you are not certain — never guess "0" just because a feature
        isn't mentioned. Absence of mention is NOT the same as absence of the feature.

Two specific traps to avoid, since they are the most common source of wrong "0" answers:
  - A document that lists "Standard Equipment" once for the whole model/range (not repeated
    per trim) means every trim covered by that document HAS those features — mark "1", not
    "U" or "0", unless a separate trim-comparison table explicitly excludes that trim.
  - In a trim-comparison table, a blank/empty cell does not automatically mean "not available".
    Only mark "0" when the cell contains an explicit negative marker (a dash, an empty circle
    next to a legend that defines it as "not available", "N/A", "-", etc.) as defined by that
    specific table's own legend. If the table's convention for a blank cell is unclear, use "U".

Only report a spec_discrepancy when the source clearly states a different value than the
current one (ignore trivial rounding/unit-notation differences). Only report a missing trim
when confident it's genuinely absent from the KNOWN TRIMS list given to you. Keep every text
field short and factual — no commentary, no explanations beyond what's requested.
"""

# Native structured output — Gemini enforces this shape server-side, so no
# markdown-fence stripping or "please return only JSON" prompting needed.
SPEC_DISCREPANCY_SCHEMA = types.Schema(
    type=types.Type.OBJECT,
    properties={
        "field": types.Schema(type=types.Type.STRING),
        "current_value": types.Schema(type=types.Type.STRING),
        "brochure_value": types.Schema(type=types.Type.STRING),
        "page_reference": types.Schema(type=types.Type.STRING),
    },
    required=["field", "current_value", "brochure_value"],
)
RESULT_ITEM_SCHEMA = types.Schema(
    type=types.Type.OBJECT,
    properties={
        "version_spec_id": types.Schema(type=types.Type.STRING),
        "brochure_covers_this_version": types.Schema(type=types.Type.BOOLEAN),
        "detected_brochure_model": types.Schema(type=types.Type.STRING),
        "verified_flags": types.Schema(
            type=types.Type.OBJECT,
            properties={
                spec: types.Schema(type=types.Type.STRING, enum=["0", "1", "U"])
                for spec in BINARY_SPECS
            },
            required=list(BINARY_SPECS),
        ),
        "discrepancies": types.Schema(type=types.Type.ARRAY, items=types.Schema(type=types.Type.STRING)),
        "spec_discrepancies": types.Schema(type=types.Type.ARRAY, items=SPEC_DISCREPANCY_SCHEMA),
    },
    required=["version_spec_id", "brochure_covers_this_version", "verified_flags", "discrepancies"],
)
MISSING_TRIM_SCHEMA = types.Schema(
    type=types.Type.OBJECT,
    properties={
        "trim_name": types.Schema(type=types.Type.STRING),
        "key_specs": types.Schema(type=types.Type.OBJECT, properties={
            f: types.Schema(type=types.Type.STRING) for f in (NUMERIC_SPECS + TEXT_SPECS)
        }),
        "page_reference": types.Schema(type=types.Type.STRING),
    },
    required=["trim_name"],
)
RESPONSE_SCHEMA = types.Schema(
    type=types.Type.OBJECT,
    properties={
        "results": types.Schema(type=types.Type.ARRAY, items=RESULT_ITEM_SCHEMA),
        "missing_trims_in_brochure": types.Schema(type=types.Type.ARRAY, items=MISSING_TRIM_SCHEMA),
    },
    required=["results"],
)
GEN_CONFIG = types.GenerateContentConfig(
    system_instruction=SYSTEM_INSTRUCTION,
    response_mime_type="application/json",
    response_schema=RESPONSE_SCHEMA,
    temperature=0,
)

# Used only for the no-brochure fallback: Gemini's Google Search grounding
# tool cannot be combined with response_schema/JSON mode in a single call, so
# this is a plain-text grounded search step; its output is then fed into a
# second, normal call using GEN_CONFIG above for structured extraction.
WEB_SEARCH_SYSTEM_INSTRUCTION = """
You are researching official automotive specifications. Search for the manufacturer's own
official website and official regional (UAE/GCC) brochure or spec page for the exact
make/model/trim given. Only use official manufacturer sources; a reputable regional dealer
page is acceptable only if no official manufacturer page is found. Report only what you
actually find, with the source for each fact. Never guess, estimate, or fill in a plausible-
sounding value. If you cannot find a fact from a real source, say so explicitly rather than
omitting it silently. Keep the report factual and concise -- specs and features only, no
marketing language.
"""
WEB_SEARCH_CONFIG = types.GenerateContentConfig(
    system_instruction=WEB_SEARCH_SYSTEM_INSTRUCTION,
    tools=[types.Tool(google_search=types.GoogleSearch())],
    temperature=0,
)

def call_gemini_web_search_with_retry(user_prompt: str):
    for attempt in range(MAX_RETRIES):
        if web_search_key_exhausted():
            return None, "The web-search API key has hit its daily quota — stopping web-search audits for today."
        if not daily_cap_try_consume():
            return None, f"Daily request cap ({DAILY_REQUEST_CAP}) reached — stopping for today."
        try:
            time.sleep(SLEEP_BETWEEN_REQUESTS)
            response = web_search_client.models.generate_content(model=MODEL_NAME, contents=user_prompt, config=WEB_SEARCH_CONFIG)
            return response.text, None
        except Exception as e:
            msg = str(e)
            is_daily_quota = "PerDay" in msg or "daily" in msg.lower()
            if is_daily_quota:
                mark_web_search_key_exhausted()
                return None, f"Gemini Web Search Daily Quota Error: {msg[:300]}"
            if any(code in msg for code in ("503", "UNAVAILABLE", "429", "RESOURCE_EXHAUSTED")):
                delay = parse_retry_delay_seconds(msg) or ((2 ** (attempt + 1)) + random.uniform(1, 2))
                print(f"[Rate Limit - web search] attempt {attempt + 1}/{MAX_RETRIES}, waiting {delay:.1f}s...")
                time.sleep(delay)
            else:
                return None, f"Gemini Web Search Error: {msg}"
    return None, "Gemini Web Search Error: Exceeded Retries"

def parse_retry_delay_seconds(err_msg: str):
    m = re.search(r'retryDelay["\']?\s*[:=]\s*["\']?(\d+(?:\.\d+)?)s', err_msg)
    if m:
        return float(m.group(1))
    m = re.search(r'seconds["\']?\s*[:=]\s*(\d+)', err_msg)
    if m:
        return float(m.group(1))
    return None

def call_gemini_with_retry(user_prompt: str):
    for attempt in range(MAX_RETRIES):
        if brochure_keys_exhausted():
            return None, "All brochure-audit API keys have hit their daily quota — stopping for today."
        if not daily_cap_try_consume():
            return None, f"Daily request cap ({DAILY_REQUEST_CAP}) reached — stopping for today."
        client, key_idx = get_active_brochure_client()
        try:
            time.sleep(SLEEP_BETWEEN_REQUESTS)
            response = client.models.generate_content(model=MODEL_NAME, contents=user_prompt, config=GEN_CONFIG)
            return response.text, None
        except Exception as e:
            msg = str(e)
            is_daily_quota = "PerDay" in msg or "daily" in msg.lower()
            if is_daily_quota:
                if advance_brochure_key(key_idx):
                    continue  # retry immediately on the new key, same attempt budget
                return None, f"Gemini Daily Quota Error (all brochure keys exhausted): {msg[:300]}"
            if any(code in msg for code in ("503", "UNAVAILABLE", "429", "RESOURCE_EXHAUSTED")):
                delay = parse_retry_delay_seconds(msg)
                if delay is None:
                    delay = (2 ** (attempt + 1)) + random.uniform(1, 2)
                print(f"[Rate Limit] attempt {attempt + 1}/{MAX_RETRIES}, waiting {delay:.1f}s...")
                time.sleep(delay)
            else:
                return None, f"Gemini API Error: {msg}"
    return None, "Gemini API Error: Exceeded Retries"

def normalize_trim_key(name):
    return re.sub(r'[^a-z0-9]', '', str(name).lower())

def _extract_specs_from_text(make, model, generation, source_text, group_rows, extra_prompt_note=""):
    """Shared structured-extraction step: given free text (brochure OR grounded
    web-search text) and a group of sheet rows for one vehicle, batches the
    rows and asks Gemini to re-derive specs using GEN_CONFIG's schema. Used by
    both the brochure path and the no-brochure web-search fallback so the
    extraction logic and output shape are identical either way."""
    all_results = []
    all_missing_trims = []
    all_version_names = [str(row.get(VERSION_COL, '')) for _, row in group_rows]

    for i in range(0, len(group_rows), GROUP_BATCH_SIZE):
        batch = group_rows[i:i + GROUP_BATCH_SIZE]
        versions_payload = []
        for idx, row in batch:
            current_flags = {c: int(row.get(c)) for c in BINARY_SPECS if c in df.columns and pd.notna(row.get(c))}
            current_specs = {c: str(row.get(c)) for c in (NUMERIC_SPECS + TEXT_SPECS) if c in df.columns and pd.notna(row.get(c))}
            versions_payload.append({
                "version_spec_id": str(row.get(ID_COL, idx)),
                "version_name": str(row.get(VERSION_COL, '')),
                "current_binary_flags": current_flags,
                "current_specs": current_specs
            })

        user_prompt = f"""
VEHICLE: {make} {model} ({generation})
{extra_prompt_note}VERSIONS: {json.dumps(versions_payload)}
KNOWN TRIMS: {json.dumps(all_version_names)}

SOURCE TEXT:
{source_text}
"""
        res_text, err = call_gemini_with_retry(user_prompt)

        if err or not res_text:
            for idx, row in batch:
                all_results.append({"version_spec_id": str(row.get(ID_COL, idx)), "error": err})
            continue

        try:
            parsed = json.loads(res_text)
            all_results.extend(parsed.get("results", []))
            all_missing_trims.extend(parsed.get("missing_trims_in_brochure", []))
        except Exception as parse_err:
            for idx, row in batch:
                all_results.append({"version_spec_id": str(row.get(ID_COL, idx)), "error": f"Parse Failure: {parse_err}"})

    return all_results, all_missing_trims


def audit_group_gemini(make, model, generation, pdf_text, group_rows, match_confidence='Exact'):
    confidence_note = ""
    if match_confidence == "Model-Only":
        confidence_note = (
            "\nMATCH CONFIDENCE: Model-Only. This brochure's filename did not contain the "
            "manufacturer name, so it was matched by model name/code alone. Model codes "
            "(e.g. H9, S7, T5) are sometimes reused across different brands. Before "
            "reporting any spec, first confirm from the SOURCE TEXT itself that the "
            f"manufacturer is actually '{make}' and the model is '{model}'. If the source "
            "text shows a different manufacturer than expected, set "
            "brochure_covers_this_version to false instead of reporting specs.\n"
        )
    return _extract_specs_from_text(make, model, generation, pdf_text, group_rows, confidence_note)


def audit_group_web_search(make, model, generation, group_rows):
    """No-brochure fallback: gathers real spec info via Gemini's Google Search
    grounding tool from official manufacturer sources, then runs the exact
    same structured extraction as the brochure path on that grounded text.
    Two Gemini calls per group (search + extract), both counted against the
    daily cap."""
    all_version_names = [str(row.get(VERSION_COL, '')) for _, row in group_rows]
    search_prompt = f"""
Find official specifications for this vehicle: {make} {model} ({generation}).
Trims/versions to cover: {json.dumps(all_version_names)}

For each trim, report: which of these features it has or lacks (only state ones you can
confirm from a real source): {", ".join(BINARY_SPECS)}.
Also report any of these specs you can confirm: {", ".join(NUMERIC_SPECS + TEXT_SPECS)}.

Cite the source (official manufacturer page or dealer page) for each fact. If you cannot find
reliable information for a trim, say so plainly instead of guessing.
"""
    grounded_text, err = call_gemini_web_search_with_retry(search_prompt)
    if err or not grounded_text or not grounded_text.strip():
        return [{"version_spec_id": str(row.get(ID_COL, idx)), "error": err or "Web search returned no text"}
                for idx, row in group_rows], []

    web_note = (
        "\nSOURCE: This text was gathered via web search (no manufacturer brochure PDF was "
        "available for this vehicle). Apply the same 0/1/U discipline -- if the web search text "
        "doesn't clearly confirm a feature for a specific trim, answer U, don't guess.\n"
    )
    return _extract_specs_from_text(make, model, generation, grounded_text, group_rows, web_note)


def audit_group_web_crosscheck(make, model, generation, group_rows, unclear_by_idx):
    """Targeted follow-up for brochure-audited rows: only the specific
    features the brochure left as 'U' (unclear) for each trim get sent to a
    web search, instead of re-auditing everything. Keeps the prompt small
    and avoids the web search overwriting anything the brochure already
    confirmed."""
    per_trim_targets = {}
    for idx, row in group_rows:
        v_id = str(row.get(ID_COL, idx))
        targets = sorted(unclear_by_idx.get(idx, set()))
        if targets:
            per_trim_targets[v_id] = {"version_name": str(row.get(VERSION_COL, '')), "features_to_check": targets}

    if not per_trim_targets:
        return {}

    search_prompt = f"""
Vehicle: {make} {model} ({generation}).
For each trim below, confirm ONLY the specific listed features from an official manufacturer
source (or a reputable dealer page if no official page exists). Do not report on any feature
not listed for that trim.

Some of these features may already be recorded as present in our data. Do not report a feature
as absent/removed unless you find clear, current evidence for that specific trim and model
year — a feature simply not being mentioned on a page is not evidence it was removed.

{json.dumps(per_trim_targets, indent=None)}

For each trim, state clearly which listed features it HAS and which it does NOT have. If you
cannot confirm a feature from a real source, say so plainly instead of guessing. Cite the
source for each fact.
"""
    grounded_text, err = call_gemini_web_search_with_retry(search_prompt)
    if err or not grounded_text or not grounded_text.strip():
        return {}

    # Reuse the same structured extraction as everywhere else, but scoped:
    # ask only about the union of targeted features across this group, to
    # keep the extraction call small too.
    all_targets = sorted({f for t in per_trim_targets.values() for f in t["features_to_check"]})
    web_note = (
        "\nSOURCE: Web search results, scoped ONLY to these specific features that the "
        f"original brochure did not clearly state: {', '.join(all_targets)}. For any other "
        "field, reply U regardless of what you might know -- only these features were "
        "actually researched for this pass.\n"
    )
    results, _ = _extract_specs_from_text(make, model, generation, grounded_text, group_rows, web_note)
    return {str(r.get('version_spec_id')): r for r in results if 'version_spec_id' in r}


def apply_crosscheck_result(idx, row, res, target_fields):
    """Surgically merges cross-check results into a row that was already
    fully processed by the brochure pass. Only touches fields in
    target_fields (the ones that were actually 'U' after the brochure), and
    never resets Accuracy_Status from scratch -- it appends to what the
    brochure pass already decided rather than replacing it."""
    if not res or res.get('error') or not res.get('brochure_covers_this_version', True):
        return False

    verified_flags = res.get('verified_flags', {}) or {}
    filled, row_changes = [], []
    for spec_col in target_fields:
        val = verified_flags.get(spec_col)
        if val is None or spec_col not in df.columns:
            continue
        val_str = str(val).strip().upper()
        if val_str not in ("0", "1"):
            continue  # still unclear even after web search -- leave untouched
        new_val = int(val_str)
        old_val = df.at[idx, spec_col]
        if values_differ(old_val, new_val):
            row_changes.append((spec_col, old_val, new_val))
            CHANGED_CELLS[(idx, spec_col)] = True
        df.at[idx, spec_col] = new_val
        filled.append(spec_col)

    if not filled:
        return False

    prior_source = df.at[idx, 'Audit_Source']
    df.at[idx, 'Audit_Source'] = f"{prior_source} + Web Search (partial)" if prior_source else "Web Search (partial)"
    fill_note = f"Web search confirmed {len(filled)} feature(s) not stated in brochure: " + ", ".join(
        f"{c}={n}" for c, _, n in row_changes
    )
    prior_note = df.at[idx, 'Discrepancies_Flagged']
    df.at[idx, 'Discrepancies_Flagged'] = (str(prior_note) + " | " + fill_note) if prior_note and str(prior_note).strip() else fill_note
    if row_changes and df.at[idx, 'Accuracy_Status'] == "Verified Accurate":
        df.at[idx, 'Accuracy_Status'] = "Flagged: Spec Discrepancy"
    log_event({
        "version_spec_id": str(row.get(ID_COL, idx)), "status": "crosscheck_fill",
        "changes": [{"field": c, "old": str(o), "new": str(n)} for c, o, n in row_changes],
    })
    return True


def make_new_row_from_trim(sample_row, vpath, trim):
    trim_name = trim.get("trim_name", "").strip()
    if not trim_name:
        return None
    dedup_key = (vpath, normalize_trim_key(trim_name))
    if dedup_key in _missing_trim_seen:
        return None
    _missing_trim_seen.add(dedup_key)

    new_row = {col: None for col in df.columns}
    for col in NEW_ROW_COPY_COLS:
        if col in df.columns:
            new_row[col] = sample_row.get(col)
    new_row[VERSION_COL] = trim_name
    new_row[ID_COL] = f"NEW_{uuid.uuid4().hex[:10]}"
    key_specs = trim.get("key_specs", {}) or {}
    for k, v in key_specs.items():
        if k in df.columns and v not in (None, ""):
            new_row[k] = v
    new_row['Brochure_File_Found'] = True
    new_row['Matched_Brochure_Path'] = vpath
    new_row['Accuracy_Status'] = "New: Found In Brochure, Not In Sheet — Needs Review"
    page_ref = trim.get("page_reference", "")
    new_row['Discrepancies_Flagged'] = (
        f"Auto-extracted from brochure{(' (' + page_ref + ')') if page_ref else ''}. "
        f"version_spec_id is a placeholder ({new_row[ID_COL]}) — replace with a real ID before publishing."
    )
    return new_row

# ============================ EXECUTION =====================================

def write_progress_snapshot(status, groups_done, groups_total, extra=None):
    snapshot = {
        "status": status,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "groups_done_this_run": groups_done,
        "groups_remaining": max(0, groups_total - groups_done),
        **(extra or {}),
    }
    with open(local(PROGRESS_NAME), "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)
    return snapshot

def report_unaudited_rows():
    pending = df.loc[df['Audit_Progress'] == "Pending"]
    if len(pending):
        cols = [c for c in [ID_COL, MAKE_COL, MODEL_COL, GEN_COL, VERSION_COL,
                             'Matched_Brochure_Path', 'Accuracy_Status'] if c in df.columns]
        pending[cols].to_csv(local(NOT_YET_AUDITED_NAME), index=False)
    return len(pending)

# Statuses that mean "hasn't actually been looked at by the AI yet, or the
# attempt failed" -- covers rows still queued, blocked on a data problem
# (bad make/model, ambiguous match) that only a sheet fix can resolve, AND
# any dynamic error message (quota, parse failure, exceeded retries) from a
# failed attempt. That last part matters: error messages include live
# details (the raw API error text) so they can never match a fixed set of
# literal strings -- checking by PREFIX is what makes a row that failed with
# "Gemini Daily Quota Error: 429 RESOURCE_EXHAUSTED..." correctly count as
# Pending (and get retried next run) instead of being miscounted as Audited
# just because its exact text isn't in a hardcoded list.
NOT_YET_AUDITED_EXACT = {
    "", "Flagged: Missing Brochure PDF", "Flagged: Missing Make/Model in Sheet",
    "Flagged: Ambiguous Brochure Match",
}
NOT_YET_AUDITED_PREFIXES = (
    "Gemini API Error", "Gemini Daily Quota Error", "Gemini Web Search Error",
    "Daily request cap", "Parse Failure",
)

def is_retryable_status(status) -> bool:
    if pd.isna(status) or status == "":
        return True
    s = str(status)
    if s in NOT_YET_AUDITED_EXACT:
        return True
    return any(s.startswith(p) for p in NOT_YET_AUDITED_PREFIXES)

def refresh_audit_progress():
    df['Audit_Progress'] = df['Accuracy_Status'].apply(
        lambda s: "Pending" if is_retryable_status(s) else "Audited"
    )

def upload_outputs():
    for fname in [OUTPUT_CSV_NAME, OUTPUT_XLSX_HIGHLIGHTED_NAME, DISCREPANCY_DETAIL_NAME,
                  NEW_TRIMS_NAME, AUDIT_LOG_NAME, SUMMARY_NAME, PROGRESS_NAME, NOT_YET_AUDITED_NAME]:
        path = local(fname)
        if os.path.exists(path):
            drive_utils.upload_or_update(drive, GDRIVE_ROOT_FOLDER_ID, path, remote_name=fname)
            print(f"  Uploaded {fname} to Drive.")

def save_checkpoint(status_note=""):
    """Every checkpoint save goes through here so Audit_Progress is always
    recomputed right before the CSV is written and uploaded -- otherwise a
    run that gets interrupted (timeout, crash, or a clean early stop) can
    leave Drive holding a CSV where Accuracy_Status was just updated but
    Audit_Progress still reflects an earlier state (blank/stale for those
    rows) until a run reaches the very end. This was the root cause of the
    blank Audit_Progress cells seen after a run that hit its quota mid-way.
    report_unaudited_rows() is also refreshed here so Not_Yet_Audited.csv
    stays accurate at every checkpoint, not just at the end."""
    refresh_audit_progress()
    report_unaudited_rows()
    df.to_csv(local(OUTPUT_CSV_NAME), index=False)
    upload_outputs()
    if status_note:
        print(status_note)

UNCLEAR_SPECS_BY_ROW = {}  # idx -> set of BINARY_SPECS column names that came back 'U' from the brochure

def apply_audit_result(idx, row, res, source_vpath, source_label):
    """Applies one Gemini result to one sheet row and sets bookkeeping columns.
    source_label is 'Brochure' or 'Web Search (No Brochure)' -- written to
    Audit_Source so every row's provenance is visible in the sheet, and,
    for the no-brochure case, prepended to Discrepancies_Flagged so it reads
    plainly even without checking the Audit_Source column."""
    v_id = str(row.get(ID_COL, idx))
    df.at[idx, 'Audit_Source'] = source_label
    no_brochure_note = "No brochure available — audited via AI web search (official sources). " \
        if source_label != "Brochure" else ""

    if res.get('error'):
        df.at[idx, 'Accuracy_Status'] = res['error']
        log_event({"version_spec_id": v_id, "status": "error", "detail": res['error'], "source": source_vpath})
        return

    if not res.get('brochure_covers_this_version', True):
        df.at[idx, 'Accuracy_Status'] = "Flagged: Mismatched Brochure PDF"
        log_event({"version_spec_id": v_id, "status": "mismatched_brochure", "source": source_vpath})
        return

    verified_flags = res.get('verified_flags', {}) or {}
    spec_discrepancies = res.get('spec_discrepancies', []) or []
    row_changes = []

    for spec_col, val in verified_flags.items():
        if spec_col not in BINARY_SPECS or spec_col not in df.columns:
            continue
        val_str = str(val).strip().upper()
        old_val = df.at[idx, spec_col]
        try:
            old_int = int(old_val) if pd.notna(old_val) else None
        except (ValueError, TypeError):
            old_int = None

        if val_str not in ("0", "1"):
            if val_str == "U":
                UNCLEAR_SPECS_BY_ROW.setdefault(idx, set()).add(spec_col)
            continue  # 'U' (or anything unexpected) -- source didn't clearly say either
                      # way for this trim; leave the existing value untouched rather
                      # than overwrite it with a guess.

        new_val = int(val_str)

        # SAFETY NET: never let a single pass (brochure OR web search) silently
        # flip a feature the sheet already says the car HAS down to absent.
        # This is exactly the failure mode that produced wrong
        # power_windows/power_door_locks corrections before -- a feature not
        # being mentioned FOR THIS TRIM in a multi-trim document got read as
        # "confirmed absent" instead of "not addressed here". A 1->0 change
        # now requires an independent, targeted second confirmation via the
        # cross-check pass before it's trusted -- same queue as 'U'.
        if old_int == 1 and new_val == 0:
            UNCLEAR_SPECS_BY_ROW.setdefault(idx, set()).add(spec_col)
            continue

        if values_differ(old_val, new_val):
            row_changes.append((spec_col, old_val, new_val, ''))
            CHANGED_CELLS[(idx, spec_col)] = True
        df.at[idx, spec_col] = new_val

    for sd in spec_discrepancies:
        field = sd.get('field', '')
        brochure_val = sd.get('brochure_value', '')
        page_ref = sd.get('page_reference', '')
        if not field or field not in df.columns or brochure_val in (None, ''):
            continue
        old_val = df.at[idx, field]
        if values_differ(old_val, brochure_val):
            row_changes.append((field, old_val, brochure_val, page_ref))
            CHANGED_CELLS[(idx, field)] = True
            df.at[idx, field] = brochure_val
        with _detail_lock:
            spec_discrepancy_details.append({
                ID_COL: v_id, MAKE_COL: row.get(MAKE_COL, ''), MODEL_COL: row.get(MODEL_COL, ''),
                VERSION_COL: row.get(VERSION_COL, ''), 'field': field, 'current_value': old_val,
                'brochure_value': brochure_val, 'page_reference': page_ref, 'source': source_vpath,
            })

    if row_changes:
        df.at[idx, 'Accuracy_Status'] = "Flagged: Spec Discrepancy"
        issue_parts = [f"{c}: {o} → {n}" + (f" (Page {p})" if p else "") for c, o, n, p in row_changes]
        df.at[idx, 'Discrepancies_Flagged'] = no_brochure_note + " | ".join(issue_parts)
    else:
        df.at[idx, 'Accuracy_Status'] = "Verified Accurate"
        df.at[idx, 'Discrepancies_Flagged'] = no_brochure_note + "No issues found — matches source."

    log_event({
        "version_spec_id": v_id, "status": df.at[idx, 'Accuracy_Status'],
        "changes": [{"field": c[0], "old": str(c[1]), "new": str(c[2]), "page": c[3]} for c in row_changes],
        "source": source_vpath,
    })


def run_pipeline():
    matched_df = df[(df['Brochure_File_Found'] == True) & (df['Matched_Brochure_Path'].astype(str).str.len() > 0)]
    to_process = [idx for idx, row in matched_df.iterrows() if row.get('Accuracy_Status') not in TERMINAL_STATUSES]

    grouped_by_pdf = {}
    for idx in to_process:
        vpath = df.at[idx, 'Matched_Brochure_Path']
        grouped_by_pdf.setdefault(vpath, []).append((idx, df.loc[idx]))

    total_groups = len(grouped_by_pdf)
    print(f"Total brochure groups pending: {total_groups} "
          f"(today's cap: {DAILY_REQUEST_CAP} requests, ~{DAILY_REQUEST_CAP // GROUP_BATCH_SIZE if GROUP_BATCH_SIZE else DAILY_REQUEST_CAP} "
          f"group(s) worth, depending on trims per brochure)")

    group_count = 0
    stopped_early = False
    for vpath, g_rows in grouped_by_pdf.items():
        if should_stop_brochure_pass():
            print(f"\n{stop_reason_brochure()} — stopping cleanly here. "
                  f"Tomorrow's scheduled run will continue automatically.")
            stopped_early = True
            break

        pdf_text = extract_pdf_text(vpath)
        if not pdf_text.strip():
            for idx, _ in g_rows:
                df.at[idx, 'Accuracy_Status'] = "Flagged: Image-Only or Unreadable PDF"
                df.at[idx, 'Audit_Source'] = "Brochure"
            continue

        first_row = g_rows[0][1]
        match_confidence = first_row.get('Match_Confidence', 'Exact')
        results, missing_trims = audit_group_gemini(
            first_row.get(MAKE_COL), first_row.get(MODEL_COL), first_row.get(GEN_COL), pdf_text, g_rows,
            match_confidence
        )

        for trim in missing_trims:
            new_row = make_new_row_from_trim(first_row, vpath, trim)
            if new_row is not None:
                with _new_rows_lock:
                    new_rows_from_brochures.append(new_row)

        res_map = {str(r.get('version_spec_id')): r for r in results if 'version_spec_id' in r}

        for idx, row in g_rows:
            v_id = str(row.get(ID_COL, idx))
            res = res_map.get(v_id, {})
            apply_audit_result(idx, row, res, vpath, "Brochure")

        group_count += 1
        if group_count % SAVE_EVERY_N_GROUPS == 0:
            write_progress_snapshot("in_progress", group_count, total_groups, {"last_brochure": vpath})
            save_checkpoint(f"[{group_count}/{total_groups}] groups done, checkpoint uploaded to Drive.")

    # --- No-brochure fallback: audit via web search instead of leaving these
    # rows permanently unaudited, as long as daily cap budget remains. Runs
    # BEFORE the cross-check pass below so any 1->0 downgrades or 'U' answers
    # it produces get queued for the same second-confirmation pass, not just
    # the ones from the brochure loop above. ---
    no_brochure_df = df[(df['Brochure_File_Found'] != True) & (df['Accuracy_Status'].apply(is_retryable_status))]
    grouped_by_vehicle = {}
    for idx, row in no_brochure_df.iterrows():
        key = (row.get(MAKE_COL), row.get(MODEL_COL), row.get(GEN_COL))
        grouped_by_vehicle.setdefault(key, []).append((idx, df.loc[idx]))

    total_web_groups = len(grouped_by_vehicle)
    if total_web_groups:
        print(f"\n{total_web_groups} vehicle(s) with no brochure at all — "
              f"attempting web-search fallback for as many as today's remaining cap allows.")

    web_group_count = 0
    for (make, model, gen), g_rows in grouped_by_vehicle.items():
        if should_stop_web_search_pass():
            print(f"\n{stop_reason_web_search()} — remaining no-brochure vehicles will be "
                  f"attempted on a future run.")
            stopped_early = True
            break

        results, _ = audit_group_web_search(make, model, gen, g_rows)
        res_map = {str(r.get('version_spec_id')): r for r in results if 'version_spec_id' in r}

        for idx, row in g_rows:
            v_id = str(row.get(ID_COL, idx))
            res = res_map.get(v_id, {})
            apply_audit_result(idx, row, res, f"web_search:{make}/{model}", "Web Search (No Brochure)")

        web_group_count += 1
        if web_group_count % SAVE_EVERY_N_GROUPS == 0:
            save_checkpoint(f"[web search {web_group_count}/{total_web_groups}] vehicles done, checkpoint uploaded to Drive.")

    # --- Cross-check pass: for ANY row where a feature came back 'U' (not
    # clearly stated), OR where a pass tried to flip an existing 1 down to 0
    # (the higher-risk case — see the safety net in apply_audit_result), do
    # one targeted, independent web search per vehicle to confirm just those
    # specific items before trusting them. Runs last so it covers gaps from
    # BOTH the brochure pass and the no-brochure web-search pass above. ---
    crosscheck_targets = {idx: fields for idx, fields in UNCLEAR_SPECS_BY_ROW.items() if fields}
    if crosscheck_targets:
        grouped_for_crosscheck = {}
        for idx, fields in crosscheck_targets.items():
            if idx not in df.index:
                continue
            row = df.loc[idx]
            key = (row.get(MAKE_COL), row.get(MODEL_COL), row.get(GEN_COL))
            grouped_for_crosscheck.setdefault(key, []).append((idx, row))

        print(f"\n{len(crosscheck_targets)} row(s) have features that were either unclear or "
              f"flagged for possible removal — attempting a targeted, independent web "
              f"cross-check for as many as today's remaining cap allows before trusting them.")
        crosscheck_group_count = 0
        for (make, model, gen), g_rows in grouped_for_crosscheck.items():
            if should_stop_web_search_pass():
                print(f"\n{stop_reason_web_search()} during cross-check pass — remaining "
                      f"unclear items will be attempted on a future run.")
                stopped_early = True
                break

            res_map = audit_group_web_crosscheck(make, model, gen, g_rows, crosscheck_targets)
            for idx, row in g_rows:
                v_id = str(row.get(ID_COL, idx))
                res = res_map.get(v_id)
                if res:
                    apply_crosscheck_result(idx, row, res, crosscheck_targets.get(idx, set()))

            crosscheck_group_count += 1
            if crosscheck_group_count % SAVE_EVERY_N_GROUPS == 0:
                save_checkpoint(f"[cross-check {crosscheck_group_count}/{len(grouped_for_crosscheck)}] "
                                 f"vehicles done, checkpoint uploaded to Drive.")

    refresh_audit_progress()
    unaudited_count = report_unaudited_rows()
    df.to_csv(local(OUTPUT_CSV_NAME), index=False)

    write_progress_snapshot(
        "stopped_daily_cap" if stopped_early else "complete",
        group_count, total_groups, {"rows_still_unaudited": int(unaudited_count)},
    )

    if CHANGED_CELLS or (df['Audit_Source'] == "Web Search (No Brochure)").any():
        df.to_excel(local(OUTPUT_XLSX_HIGHLIGHTED_NAME), index=False, sheet_name='Audited', engine='openpyxl')
        import openpyxl
        wb = openpyxl.load_workbook(local(OUTPUT_XLSX_HIGHLIGHTED_NAME))
        ws = wb['Audited']
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
        col_positions = {name: i + 1 for i, name in enumerate(df.columns)}
        row_positions = {label: pos for pos, label in enumerate(df.index)}

        # Yellow first: whole row, for anything audited without a brochure.
        web_search_rows = df.index[df['Audit_Source'] == "Web Search (No Brochure)"]
        for row_idx in web_search_rows:
            if row_idx not in row_positions:
                continue
            excel_row = row_positions[row_idx] + 2
            for col_pos in range(1, len(df.columns) + 1):
                ws[f'{get_column_letter(col_pos)}{excel_row}'].fill = yellow_fill

        # Red second, on top: individually corrected cells take priority over
        # the row-level yellow so a changed value is never mistaken for an
        # unchanged-but-web-sourced one.
        for (row_idx, col_name), _ in CHANGED_CELLS.items():
            if col_name not in col_positions or row_idx not in row_positions:
                continue
            excel_row = row_positions[row_idx] + 2
            excel_col_letter = get_column_letter(col_positions[col_name])
            ws[f'{excel_col_letter}{excel_row}'].fill = red_fill
        wb.save(local(OUTPUT_XLSX_HIGHLIGHTED_NAME))
        print(f"Highlighted {len(CHANGED_CELLS)} corrected cell(s) red, "
              f"{len(web_search_rows)} no-brochure row(s) yellow in {OUTPUT_XLSX_HIGHLIGHTED_NAME}")

    if spec_discrepancy_details:
        pd.DataFrame(spec_discrepancy_details).to_csv(local(DISCREPANCY_DETAIL_NAME), index=False)
    if new_rows_from_brochures:
        pd.DataFrame(new_rows_from_brochures).to_csv(local(NEW_TRIMS_NAME), index=False)

    summary_counts = df['Accuracy_Status'].value_counts(dropna=False)
    progress_counts = df['Audit_Progress'].value_counts(dropna=False)
    source_counts = df.loc[df['Audit_Source'] != "", 'Audit_Source'].value_counts(dropna=False)
    summary_lines = ["AUDIT SUMMARY", "=" * 40]
    summary_lines += [f"{p}: {c}" for p, c in progress_counts.items()]
    summary_lines += ["", "By Accuracy_Status:"] + [f"  {s or '(unaudited)'}: {c}" for s, c in summary_counts.items()]
    if len(source_counts):
        summary_lines += ["", "By Audit_Source:"] + [f"  {s}: {c}" for s, c in source_counts.items()]
    summary_text = "\n".join(summary_lines)
    with open(local(SUMMARY_NAME), "w", encoding="utf-8") as f:
        f.write(summary_text)
    print("\n" + summary_text)

    print("\nUploading final outputs to Drive...")
    upload_outputs()

    if stopped_early:
        print(f"\nStopped after {group_count}/{total_groups} groups this run (daily cap). "
              f"Tomorrow's scheduled GitHub Actions run picks up automatically.")
    elif unaudited_count:
        print(f"\n{unaudited_count} row(s) still unaudited (likely a PDF read/parse issue) — "
              f"see {NOT_YET_AUDITED_NAME}.")
    else:
        print("\nAll pending groups processed — nothing left in the queue.")

if __name__ == "__main__":
    run_pipeline()
