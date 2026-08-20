"""
Audit report script
- Reads data/products.csv (expected columns: product_name, brochure_path, official_site, ...feature columns...)
- For each product:
  - If brochure_path exists in repo, extract text from PDF or TXT/MD
  - Else if official_site provided, fetch page and extract visible text
  - If source text found, call OpenAI (if OPENAI_API_KEY set) with a tight prompt (prompts/audit_prompt.txt) asking to return only comma-separated 0/1 values for the feature columns
  - If OpenAI not available, fall back to a conservative deterministic scorer (0 if no exact keyword match)
- Produces data/report.csv and data/report.xlsx
- Rows audited from web but without brochure are highlighted light yellow in the Excel and have note "No brochure; audited from official website via AI"

Usage:
  python3 scripts/audit_report.py --input data/products.csv --out data/report.xlsx

Set environment variables:
  OPENAI_API_KEY (optional, for accurate AI scoring)

"""

import os
import csv
import argparse
import pandas as pd
import requests
from bs4 import BeautifulSoup
import tempfile
import pdfplumber
from dotenv import load_dotenv

# optional: OpenAI
try:
    import openai
except Exception:
    openai = None

load_dotenv()

LIGHT_YELLOW = 'FFFFE599'


def extract_text_from_pdf(path):
    try:
        with pdfplumber.open(path) as pdf:
            pages = [p.extract_text() or '' for p in pdf.pages]
        return '\n'.join(pages)
    except Exception:
        return ''


def extract_text_from_file(path):
    path = str(path)
    if path.lower().endswith('.pdf'):
        return extract_text_from_pdf(path)
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    except Exception:
        return ''


def fetch_text_from_url(url, timeout=10):
    try:
        r = requests.get(url, timeout=timeout, headers={'User-Agent': 'mdm-audit-bot/1.0'})
        if r.status_code != 200:
            return ''
        soup = BeautifulSoup(r.text, 'html.parser')
        # Remove scripts/styles
        for s in soup(['script', 'style', 'noscript']):
            s.decompose()
        txt = ' '.join(soup.stripped_strings)
        return txt
    except Exception:
        return ''


def call_openai(prompt, model='gpt-3.5-turbo', timeout=15):
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key or openai is None:
        return None
    openai.api_key = api_key
    try:
        resp = openai.ChatCompletion.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
            max_tokens=300,
            request_timeout=timeout,
        )
        return resp.choices[0].message.content.strip()
    except Exception:
        return None


def build_prompt(template, features, source_text, product_name):
    # Keep the prompt minimal and strict: only return comma-separated 0/1 values in the exact feature order
    feature_list = '\n'.join([f'- {f}' for f in features])
    prompt = template.replace('{PRODUCT_NAME}', product_name)
    prompt = prompt.replace('{FEATURE_LIST}', feature_list)
    prompt = prompt.replace('{SOURCE_TEXT}', source_text[:15000])
    return prompt


def parse_ai_response(text, features):
    # Expecting only comma separated 0/1 values e.g. "1,0,0,1"
    if not text:
        return None
    # remove brackets or surrounding code fences
    for ch in ['```', '\n']:
        text = text.strip()
    # Find first line that contains 0 or 1
    line = text.strip().splitlines()[0]
    # remove any non 0/1 and commas characters
    import re
    m = re.findall('[01]+(?:\s*,\s*[01]+)*', line)
    if not m:
        return None
    vals = m[0].replace(' ', '').split(',')
    if len(vals) != len(features):
        return None
    try:
        return [int(v) for v in vals]
    except Exception:
        return None


def conservative_score(features, source_text):
    # Very conservative: if a feature token appears in the source text (case-insensitive), mark 1, else 0
    scores = []
    text = source_text.lower()
    for f in features:
        key = f.lower()
        # simple token match
        scores.append(1 if key and key in text else 0)
    return scores


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', default='data/products.csv')
    parser.add_argument('--out', default='data/report.xlsx')
    parser.add_argument('--prompt', default='prompts/audit_prompt.txt')
    args = parser.parse_args()

    df = pd.read_csv(args.input, dtype=str).fillna('')

    # Identify standard columns
    std_cols = ['product_name', 'brochure_path', 'official_site']
    cols = list(df.columns)
    features = [c for c in cols if c not in std_cols]
    if not features:
        print('No feature columns detected. Please include feature columns after product_name, brochure_path, official_site.')
        return

    # Load prompt template
    try:
        with open(args.prompt, 'r', encoding='utf-8') as f:
            prompt_template = f.read()
    except Exception:
        prompt_template = "Answer with ONLY comma-separated 0 or 1 for the features in the order given in {FEATURE_LIST}. Do not add any explanation. Do not invent or guess if there is no evidence. Use only the provided SOURCE_TEXT. Output example: 1,0,0"

    output_rows = []
    ai_available = bool(os.getenv('OPENAI_API_KEY')) and openai is not None

    for idx, row in df.iterrows():
        product = row.get('product_name','').strip()
        brochure = row.get('brochure_path','').strip()
        site = row.get('official_site','').strip()
        source_text = ''
        source_used = None

        # try brochure first (local path or URL)
        if brochure:
            if brochure.startswith('http'):
                source_text = fetch_text_from_url(brochure)
            else:
                # try repo-local path
                if os.path.exists(brochure):
                    source_text = extract_text_from_file(brochure)
                else:
                    # maybe relative to repo root
                    p = os.path.join(os.getcwd(), brochure)
                    if os.path.exists(p):
                        source_text = extract_text_from_file(p)
            if source_text:
                source_used = 'brochure'

        if not source_text and site:
            source_text = fetch_text_from_url(site)
            if source_text:
                source_used = 'official_site'

        if source_text:
            # call AI (if available) with strict prompt
            prompt = build_prompt(prompt_template, features, source_text, product)
            ai_resp = None
            if ai_available:
                ai_resp = call_openai(prompt)
            scores = None
            if ai_resp:
                parsed = parse_ai_response(ai_resp, features)
                if parsed:
                    scores = parsed
            if scores is None:
                # fallback conservative
                scores = conservative_score(features, source_text)
            status = 'Audited'
        else:
            # no source -> mark pending and set zeros for features
            scores = [0]*len(features)
            status = 'Pending'
            source_used = 'none'

        notes = ''
        if source_used == 'official_site' and not brochure:
            notes = 'No brochure; audited from official website via AI'

        out = dict(row)
        for i, f in enumerate(features):
            out[f] = scores[i]
        out['audit_status'] = status
        out['audit_source'] = source_used
        out['note'] = notes
        output_rows.append(out)

    out_df = pd.DataFrame(output_rows)
    # Save CSV
    os.makedirs('data', exist_ok=True)
    out_df.to_csv('data/report.csv', index=False)

    # Save Excel with highlighting
    from openpyxl import load_workbook
    with pd.ExcelWriter(args.out, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='audit')

    # Post-process to color rows that were audited from web and have no brochure
    try:
        wb = load_workbook(args.out)
        ws = wb['audit']
        # find header index
        headers = [cell.value for cell in ws[1]]
        note_idx = headers.index('note') + 1
        for r in range(2, ws.max_row+1):
            note_val = ws.cell(row=r, column=note_idx).value
            if note_val and 'No brochure; audited from official website' in str(note_val):
                for c in range(1, ws.max_column+1):
                    ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type='solid')
        wb.save(args.out)
    except Exception:
        pass

    print('Produced data/report.csv and', args.out)

if __name__ == '__main__':
    main()
