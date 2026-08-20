#!/usr/bin/env python3
"""
web_audit_and_email.py (Gemini-web)

Updated to use Gemini for web searching/extraction of features and specs instead of Bing.
Behavior:
- For rows missing brochure info or with incomplete features, call Gemini with a strict schema asking it to search official manufacturer websites and authoritative pages, and return:
  - verified_flags: mapping of binary features -> 0 or 1
  - sources: mapping feature -> list of URLs (evidence)
  - audit_confident: boolean
- Only write 0/1 when explicit integers are returned. If audit_confident is False or outputs missing, mark Pending.
- Mark rows audited from web with audited_status="Audited", audited_source="Web-Gemini", note="No brochure — audited from web" and highlight them light yellow in XLSX.

Env vars (same as before):
- GDRIVE_ROOT_FOLDER_ID
- GEMINI_API_KEY
- GEMINI_MODEL (optional)
- WEB_AUDIT_TOP_K (optional)

This file replaces the previous web_audit_and_email.py to prefer Gemini-based web search/extraction.
"""

import os
import json
import time
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from google import genai
from google.genai import types

import drive_utils

# ------------------ Config ------------------
GDRIVE_ROOT_FOLDER_ID = os.environ['GDRIVE_ROOT_FOLDER_ID']
OUTPUT_CSV_NAME = os.environ.get('OUTPUT_CSV_NAME', 'AUDITED_Arabwheels_MDM_data.csv')
OUTPUT_XLSX_NAME = os.environ.get('OUTPUT_XLSX_NAME', 'AUDITED_Arabwheels_MDM_data_WEB_AUDIT.xlsx')

BINARY_SPECS = [
    'air_conditioner', 'power_windows', 'power_door_locks', 'power_steering',
    'anti_lock_braking', 'traction_control', 'immobilizer', 'cup_holders',
    'rear_folding_seat', 'rear_wiper', 'alloy_wheels', 'tubeless_tyres',
    'central_locking', 'remote_boot', 'steering_adjustment', 'tachometer',
    'child_safety_locks', 'fog_lights', 'defroster', 'defogger', 'am_fm_radio',
    'cassette_player', 'cd_player', 'sun_roof', 'moon_roof', 'is_imported',
    'cool_box', 'dvd_player', 'cruise_control', 'keyless_entry', 'rear_ac_vents',
    'front_speakers', 'rear_speakers', 'usb_and_auxillary_cable', 'heated_seats',
    'steering_switches', 'front_camera', 'rear_camera', 'push_start',
    'hill_start_assist_control', 'isofix_child_seat_anchors', 'vehicle_stability_control'
]

# Gemini setup
GEMINI_API_KEY = os.environ['GEMINI_API_KEY']
client = genai.Client(api_key=GEMINI_API_KEY)
MODEL_NAME = os.environ.get('GEMINI_MODEL', 'gemini-3.6-flash')

# Behavior
TOP_K_PAGES = int(os.environ.get('WEB_AUDIT_TOP_K', '5'))
SLEEP_BETWEEN = float(os.environ.get('WEB_AUDIT_SLEEP', '0.4'))
DAILY_REQUEST_CAP = int(os.environ.get('DAILY_REQUEST_CAP', '300'))

# ------------------ Gemini schema ------------------
VERIFIED_SCHEMA = types.Schema(
    type=types.Type.OBJECT,
    properties={
        'verified_flags': types.Schema(
            type=types.Type.OBJECT,
            properties={spec: types.Schema(type=types.Type.INTEGER) for spec in BINARY_SPECS},
            required=list(BINARY_SPECS),
        ),
        'sources': types.Schema(
            type=types.Type.OBJECT,
            properties={spec: types.Schema(type=types.Type.ARRAY, items=types.Schema(type=types.Type.STRING)) for spec in BINARY_SPECS},
            required=list(BINARY_SPECS),
        ),
        'audit_confident': types.Schema(type=types.Type.BOOLEAN),
    },
    required=['verified_flags', 'sources', 'audit_confident'],
)

GEN_CONFIG = types.GenerateContentConfig(
    system_instruction=(
        "You are an automotive spec auditor and web researcher. For the given vehicle (make, model, generation, version), "
        "search official manufacturer websites and other authoritative sources (press releases, spec pages, technical brochures online). "
        "For each binary feature requested, return whether the official sources clearly show it (1) or do not/ambiguous (0). \n"
        "Return a JSON object matching this schema exactly: {\"verified_flags\": {...}, \"sources\": {...}, \"audit_confident\": true|false}. \n"
        "- verified_flags: object mapping each feature name to integer 0 or 1. \n"
        "- sources: object mapping each feature name to an array of URL strings (evidence). Use official pages when available. If none, return empty array. \n"
        "- audit_confident: boolean indicating whether you are confident in the extraction. \n"
        "Do NOT add any extra fields, commentary, or text — return only the JSON matching the schema. Values must be integers 0 or 1. If ambiguous, choose 0."
    ),
    response_mime_type='application/json',
    response_schema=VERIFIED_SCHEMA,
    temperature=0,
)

# ------------------ Helpers ------------------

_daily_count = 0

def daily_cap_try_consume():
    global _daily_count
    if _daily_count >= DAILY_REQUEST_CAP:
        return False
    _daily_count += 1
    return True


def call_gemini_search(make, model, generation, version, features):
    """Ask Gemini to search the web and extract verified_flags and sources for the features."""
    if not daily_cap_try_consume():
        return None, 'Daily cap reached'
    # Build prompt content succinctly
    vehicle = f"{make} {model} {generation} {version}".strip()
    feature_list_text = '\n'.join([f"- {f}" for f in features])
    user_contents = (
        f"VEHICLE: {vehicle}\n"
        f"FEATURES:\n{feature_list_text}\n\n"
        "Task: Search official manufacturer websites and authoritative sources and return JSON as specified."
    )
    try:
        response = client.models.generate_content(model=MODEL_NAME, contents=user_contents, config=GEN_CONFIG)
        return response.text, None
    except Exception as e:
        return None, str(e)

# ------------------ Main flow ------------------


def main():
    print('Authenticating to Drive...')
    drive, _ = drive_utils.get_drive_service()

    # find the audited CSV on Drive; if not found, fall back to source
    file_id = drive_utils.find_file_id_by_name(drive, GDRIVE_ROOT_FOLDER_ID, OUTPUT_CSV_NAME)
    if not file_id:
        candidates = ['Arabwheels AE MDM data - MDM Version specs.csv', OUTPUT_CSV_NAME]
        for c in candidates:
            file_id = drive_utils.find_file_id_by_name(drive, GDRIVE_ROOT_FOLDER_ID, c)
            if file_id:
                break
        if not file_id:
            raise SystemExit('Could not find CSV on Drive. Place your CSV in the Drive root folder.')

    local_csv = './run_data/web_audit_input.csv'
    os.makedirs(os.path.dirname(local_csv) or '.', exist_ok=True)
    drive_utils.download_to_path(drive, file_id, local_csv)

    df = pd.read_csv(local_csv, low_memory=False, dtype=object)

    # ensure audit columns exist
    for col in ('audited_status', 'audited_source', 'note'):
        if col not in df.columns:
            df[col] = ''

    # Choose rows to audit: either no brochure OR any row with missing binary spec values
    # Missing means null/empty or not 0/1.
    def is_missing(val):
        return val is None or (str(val).strip() == '') or str(val).strip() not in ('0', '1')

    target_idxs = []
    for idx, row in df.iterrows():
        brochure_found = bool(row.get('Brochure_File_Found'))
        # collect binary features that are missing or invalid
        missing_features = [f for f in BINARY_SPECS if is_missing(row.get(f))]
        if (not brochure_found) or missing_features:
            target_idxs.append((idx, missing_features))

    print(f'Rows to audit (no brochure or missing features): {len(target_idxs)}')

    updated = 0
    for idx, missing_features in target_idxs:
        make = str(df.at[idx, 'manufacturer_name'] or '')
        model = str(df.at[idx, 'model_name'] or '')
        generation = str(df.at[idx, 'generation_name'] or '')
        version = str(df.at[idx, 'version_name'] or '')

        # If there are no missing features but no brochure (unlikely), audit all binary specs
        features_to_ask = missing_features if missing_features else BINARY_SPECS

        res_text, err = call_gemini_search(make, model, generation, version, features_to_ask)
        if err or not res_text:
            print(f'Gemini error for row {idx}: {err} — marking Pending')
            df.at[idx, 'audited_status'] = 'Pending'
            time.sleep(SLEEP_BETWEEN)
            continue

        try:
            parsed = json.loads(res_text)
        except Exception as e:
            print(f'Parse failure for row {idx}: {e} — marking Pending')
            df.at[idx, 'audited_status'] = 'Pending'
            time.sleep(SLEEP_BETWEEN)
            continue

        verified = parsed.get('verified_flags', {}) or {}
        sources = parsed.get('sources', {}) or {}
        audit_confident = bool(parsed.get('audit_confident'))

        any_written = False
        for spec in features_to_ask:
            if spec in verified:
                try:
                    v = int(verified[spec])
                except Exception:
                    continue
                if v in (0, 1):
                    df.at[idx, spec] = v
                    any_written = True
                    # store evidence link(s) in note or a dedicated column
                    srcs = sources.get(spec, [])
                    if srcs:
                        prev_note = str(df.at[idx, 'note'] or '')
                        add = f"{spec}: {', '.join(srcs)}"
                        df.at[idx, 'note'] = (prev_note + ' | ' + add).strip(' |') if prev_note else add

        if any_written and audit_confident:
            df.at[idx, 'audited_status'] = 'Audited'
            df.at[idx, 'audited_source'] = 'Web-Gemini'
            # mark rows that had no brochure explicitly
            if not bool(df.at[idx, 'Brochure_File_Found']):
                df.at[idx, 'note'] = (str(df.at[idx, 'note'] or '') + ' | No brochure — audited from web').strip(' |')
            updated += 1
        else:
            # If model returned values but audit_confident is False, keep Pending for manual review
            if any_written:
                df.at[idx, 'audited_status'] = 'Pending'
                df.at[idx, 'audited_source'] = 'Web-Gemini-Partially'
            else:
                df.at[idx, 'audited_status'] = 'Pending'
                df.at[idx, 'audited_source'] = ''

        time.sleep(SLEEP_BETWEEN)

    print(f'Web (Gemini) audit updated {updated} row(s).')

    # save CSV and XLSX (highlight light yellow rows which contain the 'No brochure — audited from web' text)
    out_csv_local = './run_data/' + OUTPUT_CSV_NAME
    df.to_csv(out_csv_local, index=False)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Audit'
    headers = list(df.columns)
    ws.append(headers)
    yellow = PatternFill(start_color='FFF7B0', end_color='FFF7B0', fill_type='solid')
    for _, row in df.iterrows():
        ws.append([row.get(h, '') for h in headers])
    note_col = None
    for i, h in enumerate(headers, start=1):
        if h == 'note':
            note_col = i
            break
    if note_col:
        for r_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=note_col)
            if isinstance(cell.value, str) and 'No brochure — audited from web' in cell.value:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r_idx, column=c).fill = yellow

    out_xlsx_local = './run_data/' + OUTPUT_XLSX_NAME
    wb.save(out_xlsx_local)

    # upload back to Drive
    drive_utils.upload_or_update(drive, GDRIVE_ROOT_FOLDER_ID, out_csv_local, remote_name=OUTPUT_CSV_NAME)
    drive_utils.upload_or_update(drive, GDRIVE_ROOT_FOLDER_ID, out_xlsx_local, remote_name=OUTPUT_XLSX_NAME)
    print('Uploaded updated CSV and XLSX to Drive.')


if __name__ == '__main__':
    main()
